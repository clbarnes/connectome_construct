"""
Convert the monoamine connectivity spreadsheet into an edgelist between neurone classes.
"""

from openpyxl import load_workbook
from collections import defaultdict, namedtuple
from itertools import product
from os.path import join
from extrasyn.paths import src_root, tgt_root


LATEST_DATA = join(src_root, 'monoamine_spreadsheet.xlsx')

IGNORE_RECEPTORS = ['dop-5', 'dop-6']

Edge = namedtuple('Edge', ['src', 'tgt', 'transmitter', 'receptor'])
monoamines = ['serotonin', 'dopamine', 'octopamine', 'tyramine']

ma_sht_name = 'Monoamine Expr'
rec_sht_name = 'Receptor Expr'


def val(cell):
    try:
        return cell.value.strip()
    except AttributeError as e:
        if 'NoneType' in str(e):
            return ''
        else:
            raise e


def ma_cell_sheet_to_dict(sheet, include_weak=False):
    mol_to_cell = defaultdict(set)
    for row in sheet.iter_rows('A3:C{}'.format(sheet.max_row)):
        if not row:
            continue
        if row[0].value:
            if '*' in row[2].value:
                continue
            if '?' in row[2].value and not include_weak:
                continue

            mol_to_cell[row[0].value.strip()].add(row[2].value.strip().replace('?', ''))

    mol_to_cell = {mol: sorted(cell_set) for mol, cell_set in mol_to_cell.items()}

    return mol_to_cell


def rec_cell_sheet_to_dict(sheet):
    transmitter_to_rec = defaultdict(set)
    rec_to_cell = defaultdict(set)
    for row in sheet.iter_rows('A3:C{}'.format(sheet.max_row)):
        if not row:
            continue

        if row[0].value:
            transmitter, receptor, cell = [item.value.strip() for item in row[:3]]

            transmitter_to_rec[transmitter].add(receptor)
            rec_to_cell[receptor].add(cell)

    transmitter_to_rec = {transmitter: sorted(rec_set) for transmitter, rec_set in transmitter_to_rec.items()}
    rec_to_cell = {receptor: sorted(cell_set) for receptor, cell_set in rec_to_cell.items()}

    return transmitter_to_rec, rec_to_cell


def main(include_weak):
    wb = load_workbook(LATEST_DATA)
    ma_sht = wb.get_sheet_by_name(ma_sht_name)
    ma_sht.calculate_dimension()
    rec_sht = wb.get_sheet_by_name(rec_sht_name)
    rec_sht.calculate_dimension()

    ma_expr = ma_cell_sheet_to_dict(ma_sht, include_weak=include_weak)
    ma_to_rec, rec_expr = rec_cell_sheet_to_dict(rec_sht)

    def generate_edges():
        for ma, src_nodes in ma_expr.items():
            for rec in ma_to_rec[ma]:
                if rec in IGNORE_RECEPTORS:
                    continue
                for src, tgt in product(src_nodes, rec_expr[rec]):
                    yield Edge(src, tgt, ma, rec)

    with open(join(tgt_root, 'ma_edgelist_classes{}.csv'.format('_include-weak' if include_weak else '')), 'w') as f:
        print('writing class edges: include weak is {}'.format(str(include_weak)))
        for edge in generate_edges():
            f.write('{},{},{},{}\n'.format(edge.src, edge.tgt, edge.transmitter, edge.receptor))


if __name__ == '__main__':
    for include_weak in [True, False]:
        main(include_weak)
    print('done')