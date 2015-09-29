"""
Convert the neuropeptide connectivity spreadsheet into an edge list between neuron classes
"""


from openpyxl import load_workbook
from collections import defaultdict, namedtuple
from os.path import join
from construct_extrasyn.paths import src_root, tgt_root
# from paths import src_root, tgt_root


LATEST_DATA = join(src_root, 'neuropeptide_spreadsheet.xlsx')

Edge = namedtuple('Edge', ['src', 'tgt', 'transmitter', 'receptor'])

pep_sht_name = 'Peptide Expr'
rec_sht_name = 'Receptor Expr'
mapping_sht_name = 'Peptides'


def val(cell):
    try:
        return cell.value.strip()
    except AttributeError as e:
        if 'NoneType' in str(e):
            return ''
        else:
            raise e


def peptide_expr_from_sheet(sheet):
    mol_to_cell = defaultdict(set)
    for row in sheet.iter_rows('A3:C{}'.format(sheet.get_highest_row())):
        if not row:
            continue
        if val(row[0]):
            if '?' in val(row[1]):
                continue

            peptide = val(row[0])
            if 'nlp-37' in peptide:
                peptide = 'pdf-2'

            mol_to_cell[peptide].add(val(row[1]))

    mol_to_cell = {mol: sorted(cell_set) for mol, cell_set in mol_to_cell.items()}

    return mol_to_cell


def receptor_expr_from_sheet(sheet):
    mol_to_cell = defaultdict(set)
    for row in sheet.iter_rows('A3:C{}'.format(sheet.get_highest_row())):
        if not row:
            continue
        if val(row[0]):
            if '?' in val(row[1]):
                continue

            mol = val(row[0])

            mol_to_cell[mol].add(val(row[1]))

    mol_to_cell = {mol: sorted(cell_set) for mol, cell_set in mol_to_cell.items()}

    return mol_to_cell


def ligand_mapping_from_sheet(sheet):
    pep_to_rec = defaultdict(set)
    for row in sheet.iter_rows('A3:C{}'.format(sheet.get_highest_row())):
        if not row:
            continue
        if val(row[0]) and val(row[1]):
            if '?' in val(row[1]):
                continue

            pep = val(row[0])

            if 'nlp-37' in pep:
                pep = 'pdf-2'

            for rec_cell in row[1:]:
                if not rec_cell or not val(rec_cell):
                    break

                pep_to_rec[pep].add(val(rec_cell))

    pep_to_rec = {mol: sorted(cell_set) for mol, cell_set in pep_to_rec.items()}

    return pep_to_rec


def main():
    wb = load_workbook(LATEST_DATA, read_only=True)

    pep_sht = wb.get_sheet_by_name(pep_sht_name)
    pep_sht.calculate_dimension(force=True)
    peptide_expr = peptide_expr_from_sheet(pep_sht)

    rec_sht = wb.get_sheet_by_name(rec_sht_name)
    rec_sht.calculate_dimension(force=True)
    rec_expr = receptor_expr_from_sheet(rec_sht)

    mapping_sht = wb.get_sheet_by_name(mapping_sht_name)
    mapping_sht.calculate_dimension(force=True)
    ligand_mapping = ligand_mapping_from_sheet(mapping_sht)

    def generate_edges():
        for transmitter in sorted(peptide_expr):
            for src in peptide_expr[transmitter]:
                for receptor in ligand_mapping[transmitter]:
                    for tgt in rec_expr.get(receptor, []):
                        yield Edge(src, tgt, transmitter, receptor)

    with open(join(tgt_root, 'np_edgelist_classes.csv'), 'w') as f:
        for edge in generate_edges():
            f.write('{},{},{},{}\n'.format(edge.src, edge.tgt, edge.transmitter, edge.receptor))


if __name__ == '__main__':
    main()
    print('done')